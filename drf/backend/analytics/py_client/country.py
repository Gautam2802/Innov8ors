import openpyxl
import requests

from openpyxl import Workbook, load_workbook

book = openpyxl.load_workbook('H-1B_Disclosure_Data_FY16.xlsx')

worksheet = book.active

endpoint_to_send = 'http://localhost:8000/api/country/'
countries = set()

for row in worksheet.iter_rows(min_col=13, max_col=13, values_only=True):
    for cell in row:
        name = cell  # Assuming the cell contains a name
        if name is not None and name not in countries:
            print(name)
            data_to_send = {
                "country": name,
            }
            countries.add(name)

            try:
                response = requests.post(endpoint_to_send, json=data_to_send)

            # Check if the request was successful (status code 200)
                if response.status_code == 200:
                    print("Data sent successfully!")
                else:
                    print(f"Failed to send data. Status Code: {response.status_code} - {response.text}")

            except requests.RequestException as e:
                print(f"Error: {e}")
        
        else:
            print("Failed to fetch data. Status Code:", response.status_code)




