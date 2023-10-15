import openpyxl

from openpyxl import Workbook, load_workbook

book = openpyxl.load_workbook('H-1B_Disclosure_Data_FY16.xlsx')

worksheet = book.active


employee = set()

for row in worksheet.iter_rows(min_row=2, max_row=10, min_col=20, max_col=20, values_only=True):
    for cell in row:
        name = cell  # Assuming the cell contains a name
        if name is not None and name not in employee:
            print(name)
            employee.add(name)