
import requests
import openpyxl

from openpyxl import Workbook, load_workbook

book = openpyxl.load_workbook('H-1B_Disclosure_Data_FY16.xlsx')

worksheet = book.active

# print(worksheet['A'].value)

# for row in worksheet.iter_rows(min_row=2, max_row=10, min_col=1, max_col=1, values_only=True):
#     for cell in row:
#         print(cell)


unique_names = set()

# Loop through cells from A2 to A10
for row in worksheet.iter_rows(min_row=2, max_row=10, min_col=1, max_col=1, values_only=True):
    for cell in row:
        name = cell  # Assuming the cell contains a name
        if name is not None and name not in unique_names:
            print(name)
            unique_names.add(name)


endpoint_to_send = 'http://localhost:8000/api/employer/'
employers = set()

columns_to_select = [1, 3, 5]

for row in worksheet.iter_rows(values_only=True):
    name = row[7]
    city = row[9]
    country = row[12]

    if name is not None and name not in employers:
        print(name)
        data_to_send = {
            "name": name,
            "city": city,
            "country": country,
        }
        employers.add(name)

        try:
            response = requests.post(endpoint_to_send, json=data_to_send)

            # Check if the request was successful (status code 200)
            if response.status_code == 200:
                print("Data sent successfully!")
            else:
                print(f"Failed to send data. Status Code: {response.status_code} - {response.text}")

        except requests.RequestException as e:
            print(f"Error: {e}")
        
    else:
        print("Failed to fetch data. Status Code:", response.status_code)




