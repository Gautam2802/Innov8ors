
import requests
import openpyxl

from openpyxl import Workbook, load_workbook

book = openpyxl.load_workbook('H-1B_Disclosure_Data_FY16.xlsx')

worksheet = book.active

# print(worksheet['A'].value)

# for row in worksheet.iter_rows(min_row=2, max_row=10, min_col=1, max_col=1, values_only=True):
#     for cell in row:
#         print(cell)


unique_names = set()

# Loop through cells from A2 to A10
for row in worksheet.iter_rows(min_row=2, max_row=10, min_col=1, max_col=1, values_only=True):
    for cell in row:
        name = cell  # Assuming the cell contains a name
        if name is not None and name not in unique_names:
            print(name)
            unique_names.add(name)


endpoint_to_send = 'http://localhost:8000/api/employer/'
employers = set()

columns_to_select = [1, 3, 5]

for row in worksheet.iter_rows(values_only=True):
    name = row[3]
    city = row[10]
    country = row[6]
    print(name)
    print(city)




